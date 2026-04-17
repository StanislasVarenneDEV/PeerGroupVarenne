import streamlit as st
import asyncio
import re
import io
from datetime import date
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Varenne Capital · Peer Group", page_icon="📊", layout="wide")

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500&display=swap');
  html, body, [class*="css"] { font-family: 'Montserrat', sans-serif; background: #fff; }
  .main { background: #fff; }
  .vhdr { border-bottom: 2px solid #8B2213; padding-bottom: 14px; margin-bottom: 20px; }
  .vtitle { font-size: 11px; letter-spacing: 5px; color: #8B2213; text-transform: uppercase; font-weight: 400; margin: 0; }
  .vsub { font-size: 10px; color: #aaa; letter-spacing: 2px; margin-top: 4px; }
  .sec { font-size: 9px; letter-spacing: 3px; text-transform: uppercase; color: #8B2213; margin: 28px 0 8px; padding-bottom: 6px; border-bottom: 0.5px solid #e8c8c4; }
  .leg { font-size: 11px; margin-bottom: 16px; }
  .lv { color: #8B2213; font-weight: 500; }
  .lf { color: #1A5276; font-weight: 500; }
  table { width: 100%; border-collapse: collapse; font-size: 12px; }
  thead th { font-size: 9px; letter-spacing: 1px; text-transform: uppercase; color: #8B2213; font-weight: 400; padding: 6px 8px; border-bottom: 1px solid #8B2213; text-align: right; white-space: nowrap; }
  thead th:first-child { text-align: center; width: 28px; }
  thead th:nth-child(2) { text-align: left; }
  tbody tr { border-bottom: 0.5px solid #f0e8e7; }
  tbody tr:hover { background: #fdf7f6; }
  td { padding: 7px 8px; text-align: right; font-size: 12px; color: #2c2c2c; white-space: nowrap; }
  td:first-child { text-align: center; color: #ccc; font-size: 11px; }
  td:nth-child(2) { text-align: left; }
  td:last-child { color: #bbb; font-size: 10px; letter-spacing: 0.5px; }
  .pos { color: #1E6B1E; font-weight: 500; }
  .neg { color: #8B2213; font-weight: 500; }
  .bld { font-weight: 600; }
  .dsh { color: #ddd; }
  .rv td:nth-child(2) { color: #8B2213; font-weight: 600; }
  .rv td:first-child { border-left: 3px solid #8B2213; }
  .rf td:nth-child(2) { color: #1A5276; }
  .rf td:first-child { border-left: 3px solid #1A5276; }
  .stButton button { background: #8B2213 !important; color: #fff !important; border: none !important; border-radius: 2px !important; font-size: 11px !important; letter-spacing: 2px !important; text-transform: uppercase !important; padding: 10px 24px !important; }
  .stDownloadButton button { background: #fff !important; color: #8B2213 !important; border: 1px solid #8B2213 !important; border-radius: 2px !important; font-size: 11px !important; letter-spacing: 2px !important; text-transform: uppercase !important; }
  .foot { font-size: 9px; color: #ccc; margin-top: 36px; padding-top: 10px; border-top: 0.5px solid #f0e0de; }
</style>
""", unsafe_allow_html=True)

# ── Données ──

VARENNE = {"Varenne Long Short - P","Varenne Long Short - A","Varenne Global - P","Varenne Global - A","Varenne Conviction - P","Varenne Conviction - A","Varenne Valeur - P","Varenne Valeur - A"}
FOCUS   = {"Moneta Long Short","JPM Europe Absolute Equity Alpha","Eleva Absolute Return","Eurose","Moneta Multicaps"}

FALLBACK = {
    "Carmignac Long Short": ("carmignac","https://www.carmignac.com/fr-fr/nos-fonds-notre-gestion/carmignac-portfolio-long-short-european-equities-LU1317704051-a-eur-acc/performances#fundpage-start"),
}

FONDS = {
    "Fonds Long — Short": [
        ("BDL Rempart",                      "13406",   0.0422,  0.1798, -0.0281,  0.0600,  0.1490,  0.1610, -0.0640,  0.5074, "FR0010174144"),
        ("Varenne Long Short - P",           "2115146", 0.0378,  0.0811,  0.0600,   None,    None,    None,    None,    None,   "LU2722195240"),
        ("Pictet Atlas - Titan",             "918986",  0.0376,  0.1263,  0.1439,  0.0633, -0.0938,  0.1471,  None,    0.3732, "LU2206556016"),
        ("Varenne Long Short - A",           "2115144", 0.0360,  0.0730,  0.0473,   None,    None,    None,    None,    None,   "LU2722194946"),
        ("Moneta Long Short",                "52199",   None,    0.2275,  0.0280,  0.0600, -0.0320,  0.0620,  0.0240,  None,   "FR0010400762"),
        ("Pictet Atlas",                     "435720",  0.0194,  0.0669,  0.0851,  0.0373, -0.0564,  0.0617,  0.1078,  0.1838, "LU1433332854"),
        ("JPM Europe Absolute Equity Alpha", "100724",  0.0187,  0.1120,  0.1875,  0.0200,  0.0740,  0.1540, -0.0370,  0.5856, "LU1001748711"),
        ("Eleva Absolute Return Dynamic",    "2117653", 0.0080,  0.1173,   None,    None,    None,    None,    None,    None,   "LU2719142965"),
        ("Eleva Absolute Return",            "733949",  0.0043,  0.0491,  0.0745,  0.0454, -0.0237,  0.0541,  0.0781,  0.2068, "LU1920211973"),
        ("Carmignac Long Short",             "1317704", None,    0.0778,  0.1741,  0.0007, -0.0633,  0.1295,  0.0692,  None,   "LU1317704051"),
    ],
    "Fonds Long Only Monde": [
        ("Varenne Global - P",               "2023975", 0.0622,  0.1401,  0.1112,  0.2179, -0.2549,  0.2836,  0.1233,  0.4110, "LU2358391725"),
        ("Varenne Conviction - P",           "2115142", 0.0625,  0.1442,  0.1303,  0.2588, -0.2653,  0.3615,  None,    None,   "LU2722194516"),
        ("Varenne Global - A",               "2023974", 0.0605,  0.1361,  0.1053,  0.2095, -0.2601,  0.2758,  0.1096,  0.3710, "LU2358389745"),
        ("Varenne Conviction - A",           "2115147", 0.0600,  0.1359,  0.1185,  0.2588, -0.2653,  0.3615,  None,    None,   "LU2722195596"),
        ("Carmignac Investissement",         "9364",    0.0406,  0.1745,  0.2503,  0.1892, -0.1833,  0.0397,  0.3365,  0.4362, "FR0010148981"),
        ("Echiquier World Equity Growth",    "63249",   0.0063,  0.0337,  0.2078,  0.1800, -0.1620,  0.1080,  0.1640,  0.3122, "FR0010859769"),
        ("Comgest Monde - C",                "15111",  -0.0322, -0.0042,  0.1539,  0.2184, -0.2009,  0.1598,  0.1160,  0.1968, "FR0000284689"),
    ],
    "Fonds utilisés par les CGP face à Valeur": [
        ("Varenne Valeur - P",               "2028844", 0.0392,  0.0962,  0.0694,  0.1280, -0.1584,  0.2016,  0.1108,  0.2825, "LU2358390321"),
        ("Varenne Valeur - A",               "2026192", 0.0377,  0.0907,  0.0659,  0.1213, -0.1634,  0.1955,  0.1055,  0.2507, "LU2358392376"),
        ("Carmignac Patrimoine",             "9362",    None,    0.1212,  0.0706,  0.0220, -0.0938, -0.0088,  0.1240,  None,   "FR0010135103"),
        ("Eurose",                           "6875",    0.0143,  0.0793,  0.0300,  0.0860, -0.0309,  0.0717, -0.0426,  0.2230, "FR0007051040"),
        ("Keren Patrimoine",                 "6865",    0.0170,  0.0777,  0.0668,  0.1098,  0.0964,  0.1086, -0.0280,  0.2355, "FR0000980427"),
        ("Ruffer Total Return",              "85327",   0.0108,  0.0891, -0.0287, -0.0835,  0.0405,  0.0848,  0.1192,  0.0364, "LU0638558717"),
        ("Moneta Multicaps",                 "12206",   0.0166,  0.2619, -0.0259,  0.0627, -0.0825,  0.2122,  0.0650,  0.3469, "FR0010298596"),
        ("DNCA Evolutif",                    "110033", -0.0223,  0.0549,  0.1402,  0.1578, -0.1538,  0.1440,  0.0303,  0.2483, "LU1055118761"),
        ("Rco Valor",                        "84944",  -0.0097,  0.1579,  0.1630,  0.1260, -0.0838,  0.1232,  0.0635,  0.3909, "FR1011261197"),
        ("Sextant Grand Large",              "19436",  -0.0188,  0.0286,  0.0118,  0.0920, -0.0530,  0.0380, -0.0180,  0.0573, "FR0010286013"),
        ("Flossbach Multi Opportunities",    "132411", -0.0280,  0.0217,  0.0891,  0.0865, -0.1340,  0.1070,  0.0334,  0.0684, "LU1245469744"),
        ("JPM Global Macro",                 "9964",   -0.0357,  0.0571, -0.0229, -0.0245, -0.1371,  0.0314,  0.1083, -0.1619, "LU0115098948"),
        ("Acatis Value Event",               "76262",  -0.0370,  0.0053,  0.0990,  0.1462, -0.1244,  0.1383,  0.0708,  0.1576, "DE000A0X7541"),
    ],
}

# ── Scraping ──

def ppct(t):
    try: return float(re.sub(r'[^\d.,-]','',t).replace(',','.')) / 100
    except: return None

def exq(html):
    soup = BeautifulSoup(html,"lxml")
    lines = [l.strip() for l in soup.get_text().split('\n') if l.strip()]
    ytd = ans5 = None
    for i,line in enumerate(lines):
        if "1er janvier" in line and ytd is None:
            for j in range(1,4):
                if i+j<len(lines) and "%" in lines[i+j]: ytd=ppct(lines[i+j]); break
        if "Perf. 5 ans" in line and ans5 is None:
            for j in range(1,4):
                if i+j<len(lines) and "%" in lines[i+j]: ans5=ppct(lines[i+j]); break
        if ytd and ans5: break
    return ytd, ans5

def exc(html):
    text = BeautifulSoup(html,"lxml").get_text()
    idx = text.find("Depuis le début de l'année")
    if idx == -1: return None, None
    m = re.findall(r'[+-]?\d+\.?\d*\s*%', text[idx:idx+500])
    return (ppct(m[0]) if m else None), (ppct(m[5]) if len(m)>5 else None)

async def scrape(cb):
    res = {}
    all_f = [(f[0],f[1]) for g in FONDS.values() for f in g]
    async with async_playwright() as p:
        br = await p.chromium.launch(headless=True)
        ctx = await br.new_context(locale="fr-FR")
        for i,(nom,fid) in enumerate(all_f):
            cb(i/len(all_f), f"Scraping {nom}...")
            ytd = ans5 = None
            try:
                pg = await ctx.new_page()
                await pg.goto(f"https://www.quantalys.com/Fonds/{fid}", wait_until="networkidle", timeout=30000)
                ytd, ans5 = exq(await pg.content())
                await pg.close()
            except: pass
            if (ytd is None or ans5 is None) and nom in FALLBACK:
                src, url = FALLBACK[nom]
                try:
                    pg = await ctx.new_page()
                    await pg.goto(url, wait_until="networkidle", timeout=40000)
                    await pg.wait_for_timeout(3000)
                    y2,a2 = exc(await pg.content()) if src=="carmignac" else (None,None)
                    await pg.close()
                    if ytd is None: ytd=y2
                    if ans5 is None: ans5=a2
                except: pass
            res[nom] = {"ytd":ytd,"5ans":ans5}
            await asyncio.sleep(0.3)
        await br.close()
    cb(1.0,"Terminé ✓")
    return res

# ── HTML tableau ──

def fmt(v, bold=False):
    if v is None: return '<span class="dsh">—</span>'
    c = ("pos" if v>0 else "neg" if v<0 else "") + (" bld" if bold else "")
    return f'<span class="{c.strip()}">{v*100:+.2f}%</span>'

def table_html(groupe, liste, scraped, today):
    h = f'<p class="sec">{groupe}</p>'
    h += f'<table><thead><tr><th>#</th><th>Fonds</th><th>YTD au {today}</th><th>2025</th><th>2024</th><th>2023</th><th>2022</th><th>2021</th><th>2020</th><th>Perf. 5 ans</th><th>ISIN</th></tr></thead><tbody>'
    for rang,f in enumerate(liste,1):
        nom,fid,yh,p25,p24,p23,p22,p21,p20,p5h,isin = f
        s = scraped.get(nom,{})
        ytd = s.get("ytd") if s.get("ytd") is not None else yh
        p5  = s.get("5ans") if s.get("5ans") is not None else p5h
        rc = "rv" if nom in VARENNE else ("rf" if nom in FOCUS else "")
        h += f'<tr class="{rc}"><td>{rang}</td><td>{nom}</td><td>{fmt(ytd,True)}</td>'
        for v in [p25,p24,p23,p22,p21,p20]: h += f'<td>{fmt(v)}</td>'
        h += f'<td>{fmt(p5,True)}</td><td>{isin}</td></tr>'
    return h + '</tbody></table>'

# ── Export Excel ──

def gen_excel(scraped, today):
    R="8B2213";B="1A5276";W="FFFFFF";GL="F7F7F7";GT="2C2C2C";GI="BBBBBB";VP="1E6B1E";SEP="ECECEC"
    fp=lambda c:PatternFill("solid",start_color=c,fgColor=c)
    brd=lambda l=None:Border(left=Side(style="thick",color=l) if l else Side(style=None),bottom=Side(style="hair",color=SEP))
    def sc(ws,r,c,v="",**kw):
        cell=ws.cell(row=r,column=c,value=v)
        for k,val in kw.items(): setattr(cell,k,val)
        return cell
    def pc(ws,r,c,val,bold=False,bg=W,left=None):
        cell=ws.cell(row=r,column=c); cell.fill=fp(bg); cell.border=brd(left if c==1 else None)
        if val is not None:
            cell.value=val; cell.number_format="0.00%"
            cell.alignment=Alignment(horizontal="right",vertical="center")
            cell.font=Font(name="Calibri",size=9,bold=bold,color=VP if val>0 else (R if val<0 else GT))
        else:
            cell.value="—"; cell.alignment=Alignment(horizontal="center",vertical="center")
            cell.font=Font(name="Calibri",size=9,color="DDDDDD")

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Peers"; ws.sheet_view.showGridLines=False
    COLS=["#","Fonds",f"YTD au {today}","2025","2024","2023","2022","2021","2020","Perf. 5 ans","ISIN"]
    WIDTHS=[4,36,14,9,9,9,9,9,9,12,16]; NB=len(COLS)

    ws.row_dimensions[1].height=32
    for c in range(1,NB+1): ws.cell(row=1,column=c).fill=fp(W)
    sc(ws,1,1,"VARENNE CAPITAL  ·  Peer Group",font=Font(name="Calibri",size=14,color=R),fill=fp(W),alignment=Alignment(horizontal="left",vertical="center"))
    ws.row_dimensions[2].height=2
    for c in range(1,NB+1): ws.cell(row=2,column=c).fill=fp(R)
    ws.row_dimensions[3].height=18
    for c in range(1,NB+1): ws.cell(row=3,column=c).fill=fp(W)
    sc(ws,3,1,"▌ Fonds Varenne Capital",font=Font(name="Calibri",size=8,color=R,bold=True),fill=fp(W),alignment=Alignment(horizontal="left",vertical="center"))
    sc(ws,3,4,"▌ Fonds Focus Europe",font=Font(name="Calibri",size=8,color=B,bold=True),fill=fp(W),alignment=Alignment(horizontal="left",vertical="center"))

    cr=4
    for groupe,liste in FONDS.items():
        cr+=1
        for c in range(1,NB+1):
            ws.cell(row=cr,column=c).fill=fp(W)
            ws.cell(row=cr,column=c).border=Border(bottom=Side(style="thin",color="E0D0CE"))
        sc(ws,cr,1,groupe.upper(),font=Font(name="Calibri",size=8,color=R),fill=fp(W),alignment=Alignment(horizontal="left",vertical="bottom"),border=Border(bottom=Side(style="thin",color="E0D0CE"))); cr+=1
        ws.row_dimensions[cr].height=20
        for ci,h in enumerate(COLS,1):
            ah="right" if ci>2 else("center" if ci==1 else "left")
            sc(ws,cr,ci,h,font=Font(name="Calibri",size=8,color="888888"),fill=fp(W),alignment=Alignment(horizontal=ah,vertical="center"),border=Border(bottom=Side(style="thin",color="CCCCCC"))); cr+=1
        for rang,f in enumerate(liste,1):
            nom,fid,yh,p25,p24,p23,p22,p21,p20,p5h,isin=f
            s=scraped.get(nom,{}); ytd=s.get("ytd") if s.get("ytd") is not None else yh; p5=s.get("5ans") if s.get("5ans") is not None else p5h
            iv=nom in VARENNE; inf=nom in FOCUS; left=R if iv else(B if inf else None); bg=W if rang%2!=0 else GL
            ws.row_dimensions[cr].height=15
            sc(ws,cr,1,rang,font=Font(name="Calibri",size=8,color="CCCCCC"),fill=fp(bg),alignment=Alignment(horizontal="center",vertical="center"),border=brd(left))
            sc(ws,cr,2,nom,font=Font(name="Calibri",size=9,bold=iv,color=R if iv else(B if inf else GT)),fill=fp(bg),alignment=Alignment(horizontal="left",vertical="center"),border=brd())
            pc(ws,cr,3,ytd,bold=True,bg=bg)
            for ci,val in zip(range(4,10),[p25,p24,p23,p22,p21,p20]): pc(ws,cr,ci,val,bg=bg)
            pc(ws,cr,10,p5,bold=True,bg=bg)
            sc(ws,cr,11,isin,font=Font(name="Calibri",size=8,color=GI),fill=fp(bg),alignment=Alignment(horizontal="center",vertical="center"),border=brd()); cr+=1

    for i,w in enumerate(WIDTHS,1): ws.column_dimensions[get_column_letter(i)].width=w
    cr+=2
    sc(ws,cr,1,f"Source : Quantalys / Carmignac.com  ·  Performances en EUR  ·  {today}  ·  Les performances passées ne préjugent pas des performances futures",font=Font(name="Calibri",size=7,color="CCCCCC",italic=True),alignment=Alignment(horizontal="left",vertical="center"))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── Interface ──

today = date.today().strftime("%d/%m/%Y")

st.markdown(f"""
<div class="vhdr">
  <p class="vtitle">Varenne Capital</p>
  <p class="vsub">Peer Group · Suivi des performances</p>
</div>
<div class="leg">
  <span class="lv">▌ Fonds Varenne Capital</span>&nbsp;&nbsp;&nbsp;
  <span class="lf">▌ Fonds Focus Europe</span>
</div>
""", unsafe_allow_html=True)

if "scraped" not in st.session_state:
    st.session_state.scraped = {f[0]:{"ytd":f[2],"5ans":f[9]} for g in FONDS.values() for f in g}
if "last_update" not in st.session_state:
    st.session_state.last_update = "Données statiques — cliquez Mettre à jour"

c1, c2, c3 = st.columns([1,1,5])
with c1: update = st.button("Mettre à jour")
with c2:
    st.download_button(
        "Télécharger Excel",
        data=gen_excel(st.session_state.scraped, today),
        file_name=f"PeerGroup_Varenne_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if update:
    bar=st.progress(0); info=st.empty()
    def cb(v,m): bar.progress(v); info.text(m)
    with st.spinner("Scraping en cours (~3 min)..."):
        st.session_state.scraped = asyncio.run(scrape(cb))
        st.session_state.last_update = f"Mis à jour le {today}"
    bar.empty(); info.empty()
    st.success(f"✓ Données mises à jour — {today}")

for groupe, liste in FONDS.items():
    st.markdown(table_html(groupe, liste, st.session_state.scraped, today), unsafe_allow_html=True)

st.markdown(f'<div class="foot">Source : Quantalys · Carmignac.com · Performances en EUR · {st.session_state.last_update} · Les performances passées ne préjugent pas des performances futures</div>', unsafe_allow_html=True)
